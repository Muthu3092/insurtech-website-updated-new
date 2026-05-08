"""Sales CRM router — Leads, Deals, Lead-Deal Linkages, Activities, Tasks,
AI Calls, AI Agents, WhatsApp, Meetings, Lookup. Replicates Sales-hub-CRM functionality
adapted for the Insurance Tech Platform (Customer Name + Passport + IC fields)."""
import io
import uuid
import math
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from openpyxl import load_workbook

from auth import get_current_user, require_roles
from database import db
from services.ai_services import (
    initiate_elevenlabs_call,
    get_elevenlabs_conversation,
    download_elevenlabs_audio,
    send_whatsapp,
    send_meeting_invite_email,
    build_ics,
    compute_ai_lead_score,
)

router = APIRouter(tags=["sales-crm"])


def _id() -> str:
    return uuid.uuid4().hex


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


# ============ MODELS ============
class LeadCreate(BaseModel):
    # Customer / Company info (renamed from clinic; "name" still holds customer-company name)
    name: str = ""               # Customer Name (Company / Org / Individual)
    country: str = "Malaysia"
    state: str = ""
    city: str = ""
    postcode: str = ""
    company_size: str = ""
    address: str = ""
    website: str = ""
    industry: str = ""
    # PIC (Person In Charge)
    pic_name: str = ""
    title: str = ""
    # Identity
    ic_number: str = ""           # NEW: Malaysian IC
    passport_number: str = ""     # NEW: Passport
    date_of_birth: str = ""       # YYYY-MM-DD
    gender: str = ""              # male / female / other
    race: str = ""                # Malay / Chinese / Indian / Other
    # Contact
    phone: str = ""               # Mobile number
    office_number: str = ""       # Phone (office / landline)
    fax_number: str = ""
    email: str = ""
    linkedin: str = ""
    # Pipeline
    pipeline_status: str = "new"
    status: str = "new"
    source: str = ""
    notes: str = ""
    is_public: bool = False
    company: str = ""  # legacy alias kept for back-compat


class LeadUpdateModel(LeadCreate):
    converted_to_client: Optional[bool] = None
    converted_to_customer: Optional[bool] = None


class ActivityCreate(BaseModel):
    type: str
    description: str
    notes: Optional[str] = ""
    scheduled_at: Optional[str] = None


class DealCreate(BaseModel):
    title: str
    value: float = 0.0
    stage: str = "lead"
    expected_close_date: Optional[str] = None
    notes: Optional[str] = ""
    knowledge_base_content: Optional[str] = None
    linked_company_ids: List[str] = []


class DealUpdate(BaseModel):
    title: Optional[str] = None
    value: Optional[float] = None
    stage: Optional[str] = None
    expected_close_date: Optional[str] = None
    notes: Optional[str] = None
    knowledge_base_content: Optional[str] = None
    linked_company_ids: Optional[List[str]] = None


class LinkageCreate(BaseModel):
    lead_id: str
    deal_id: str
    pipeline_status: str = "new"
    notes: Optional[str] = ""


class TaskCreate(BaseModel):
    title: str
    description: Optional[str] = ""
    lead_id: Optional[str] = None
    deal_id: Optional[str] = None
    status: str = "pending"
    priority: str = "medium"
    payment_status: str = "unpaid"
    pipeline_status: Optional[str] = None
    due_date: Optional[str] = None


class AiAgentCreate(BaseModel):
    name: str
    agent_id: str
    description: Optional[str] = ""
    phone_number_id: Optional[str] = None
    voice_id: Optional[str] = None


class AiCallInitiate(BaseModel):
    lead_id: str
    deal_id: str
    agent_name: Optional[str] = None
    agent_id: Optional[str] = None
    phone: Optional[str] = None
    use_dynamic_selection: bool = False
    call_purpose: str = "follow_up"


class WhatsAppSend(BaseModel):
    contact_id: Optional[str] = None
    phone: str
    message: str


class MeetingSchedule(BaseModel):
    lead_id: str
    title: str
    description: Optional[str] = ""
    meeting_type: str = "online"
    date: str
    start_time: str = "10:00"
    duration_minutes: int = 30
    location: Optional[str] = ""
    send_invite: bool = True


class ConvertToCustomer(BaseModel):
    services: List[Dict[str, Any]] = []
    notes: Optional[str] = ""


# ============ HELPERS ============
def _strip_id(d: Dict) -> Dict:
    if d:
        d.pop("_id", None)
    return d


# ============ LEADS ============
@router.post("/leads")
async def create_lead(body: LeadCreate, user: dict = Depends(require_roles("admin", "agent"))):
    if not body.name:
        raise HTTPException(400, "Customer Name is required")
    doc = body.model_dump()
    doc["company"] = doc.get("company") or doc["name"]
    doc.update({
        "id": _id(),
        "owner_id": user["id"],
        "owner_name": user.get("full_name", ""),
        "ai_score": await compute_ai_lead_score(body.model_dump()),
        "converted_to_client": False,
        "converted_to_customer": False,
        "created_at": _now(),
        "updated_at": _now(),
    })
    await db.leads.insert_one(doc)
    await db.activities.insert_one({
        "id": _id(), "entity_type": "lead", "entity_id": doc["id"],
        "type": "import", "action": "create",
        "description": f"Lead created: {doc['name']}",
        "notes": "", "user_id": user["id"], "user_name": user.get("full_name", ""),
        "created_at": _now(),
    })
    return _strip_id(doc)


@router.get("/leads")
async def list_leads(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=200),
    pipeline_status: Optional[str] = None,
    state: Optional[str] = None,
    search: Optional[str] = None,
    _: dict = Depends(require_roles("admin", "agent")),
):
    q: Dict[str, Any] = {}
    if pipeline_status and pipeline_status != "all":
        q["pipeline_status"] = pipeline_status
    if state and state != "all":
        q["state"] = state
    if search:
        q["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"pic_name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"company": {"$regex": search, "$options": "i"}},
        ]
    total = await db.leads.count_documents(q)
    cursor = db.leads.find(q, {"_id": 0}).sort("created_at", -1).skip((page - 1) * limit).limit(limit)
    items = await cursor.to_list(limit)
    return {
        "items": items, "total": total,
        "page": page, "limit": limit,
        "total_pages": math.ceil(total / limit) if total else 0,
    }


@router.get("/leads/{lead_id}")
async def get_lead(lead_id: str, _: dict = Depends(require_roles("admin", "agent"))):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    return lead


@router.put("/leads/{lead_id}")
async def update_lead(lead_id: str, body: LeadUpdateModel, user: dict = Depends(require_roles("admin", "agent"))):
    updates = {k: v for k, v in body.model_dump(exclude_unset=True).items() if v is not None}
    if "name" in updates and updates["name"]:
        updates["company"] = updates["name"]
    updates["updated_at"] = _now()
    if any(k in updates for k in ("phone", "email", "name", "pipeline_status")):
        # recompute score
        merged = await db.leads.find_one({"id": lead_id}, {"_id": 0}) or {}
        merged.update(updates)
        updates["ai_score"] = await compute_ai_lead_score(merged)
    await db.leads.update_one({"id": lead_id}, {"$set": updates})
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    return lead


@router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str, _: dict = Depends(require_roles("admin", "agent"))):
    await db.leads.delete_one({"id": lead_id})
    await db.activities.delete_many({"entity_type": "lead", "entity_id": lead_id})
    return {"deleted": True}


@router.post("/leads/{lead_id}/refresh-score")
async def refresh_score(lead_id: str, _: dict = Depends(require_roles("admin", "agent"))):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    score = await compute_ai_lead_score(lead)
    await db.leads.update_one({"id": lead_id}, {"$set": {"ai_score": score, "updated_at": _now()}})
    return {"ai_score": score}


@router.post("/leads/{lead_id}/convert")
async def convert_lead(lead_id: str, body: ConvertToCustomer, user: dict = Depends(require_roles("admin", "agent"))):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    customer = {
        "id": _id(),
        "name": lead.get("name"),
        "pic_name": lead.get("pic_name"),
        "email": lead.get("email"),
        "phone": lead.get("phone"),
        "ic_number": lead.get("ic_number"),
        "passport_number": lead.get("passport_number"),
        "address": lead.get("address"),
        "city": lead.get("city"),
        "state": lead.get("state"),
        "country": lead.get("country"),
        "services": body.services,
        "notes": body.notes,
        "source_lead_id": lead_id,
        "created_at": _now(),
    }
    await db.crm_customers.insert_one(customer)
    await db.leads.update_one({"id": lead_id}, {"$set": {
        "converted_to_client": True, "converted_to_customer": True,
        "status": "qualified", "updated_at": _now(),
    }})
    return _strip_id(customer)


@router.post("/leads/import")
async def import_leads(file: UploadFile = File(...), user: dict = Depends(require_roles("admin", "agent"))):
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Please upload an Excel file (.xlsx or .xls)")
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(400, f"Failed to parse file: {e}")
    if not rows or len(rows) < 2:
        raise HTTPException(400, "File appears empty")
    header = [str(c or "").strip().lower().replace(" ", "_") for c in rows[0]]
    imported = 0
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        rec = {h: (str(v).strip() if v is not None else "") for h, v in zip(header, row)}
        name = rec.get("name") or rec.get("customer_name") or rec.get("company") or rec.get("clinic_name")
        if not name:
            continue
        doc = {
            "id": _id(),
            "name": name,
            "company": name,
            "pic_name": rec.get("pic_name") or rec.get("contact_name") or "",
            "title": rec.get("title") or rec.get("job_title") or "",
            "email": rec.get("email") or "",
            "phone": rec.get("phone") or rec.get("mobile") or "",
            "office_number": rec.get("office_number") or rec.get("office") or "",
            "ic_number": rec.get("ic_number") or rec.get("ic") or "",
            "passport_number": rec.get("passport_number") or rec.get("passport") or "",
            "industry": rec.get("industry") or "",
            "company_size": rec.get("company_size") or rec.get("size") or "",
            "country": rec.get("country") or "Malaysia",
            "state": rec.get("state") or "",
            "city": rec.get("city") or "",
            "postcode": rec.get("postcode") or rec.get("zip") or "",
            "address": rec.get("address") or "",
            "website": rec.get("website") or "",
            "source": rec.get("source") or "import",
            "notes": rec.get("notes") or "",
            "pipeline_status": "new",
            "status": "new",
            "owner_id": user["id"], "owner_name": user.get("full_name", ""),
            "converted_to_client": False, "converted_to_customer": False,
            "created_at": _now(), "updated_at": _now(),
        }
        doc["ai_score"] = await compute_ai_lead_score(doc)
        await db.leads.insert_one(doc)
        imported += 1
    return {"imported_count": imported}


# ============ ACTIVITIES ============
@router.get("/leads/{lead_id}/activities")
async def list_activities(lead_id: str, _: dict = Depends(require_roles("admin", "agent"))):
    items = await db.activities.find(
        {"entity_type": "lead", "entity_id": lead_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    return {"activities": items}


@router.post("/leads/{lead_id}/activities")
async def add_activity(lead_id: str, body: ActivityCreate, user: dict = Depends(require_roles("admin", "agent"))):
    doc = {
        "id": _id(), "entity_type": "lead", "entity_id": lead_id,
        "type": body.type, "action": body.type,
        "description": body.description, "notes": body.notes or "",
        "scheduled_at": body.scheduled_at,
        "user_id": user["id"], "user_name": user.get("full_name", ""),
        "created_at": _now(),
    }
    await db.activities.insert_one(doc)
    return _strip_id(doc)


# ============ DEALS ============
@router.post("/deals")
async def create_deal(body: DealCreate, user: dict = Depends(require_roles("admin", "agent"))):
    doc = body.model_dump()
    doc.update({"id": _id(), "owner_id": user["id"], "created_at": _now(), "updated_at": _now()})
    await db.deals.insert_one(doc)
    return _strip_id(doc)


@router.get("/deals")
async def list_deals(_: dict = Depends(require_roles("admin", "agent"))):
    items = await db.deals.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items


@router.get("/deals/{deal_id}")
async def get_deal(deal_id: str, _: dict = Depends(require_roles("admin", "agent"))):
    d = await db.deals.find_one({"id": deal_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Deal not found")
    return d


@router.put("/deals/{deal_id}")
async def update_deal(deal_id: str, body: DealUpdate, _: dict = Depends(require_roles("admin", "agent"))):
    updates = {k: v for k, v in body.model_dump(exclude_unset=True).items() if v is not None}
    updates["updated_at"] = _now()
    await db.deals.update_one({"id": deal_id}, {"$set": updates})
    return await db.deals.find_one({"id": deal_id}, {"_id": 0})


@router.delete("/deals/{deal_id}")
async def delete_deal(deal_id: str, _: dict = Depends(require_roles("admin", "agent"))):
    await db.deals.delete_one({"id": deal_id})
    return {"deleted": True}


@router.get("/deals/{deal_id}/agents")
async def deal_agents(deal_id: str, _: dict = Depends(require_roles("admin", "agent"))):
    deal = await db.deals.find_one({"id": deal_id}, {"_id": 0})
    if not deal:
        return {"agents": [], "selection_mode": "manual"}
    agent_ids = deal.get("ai_agent_ids", [])
    agents = await db.ai_agents.find({"id": {"$in": agent_ids}}, {"_id": 0}).to_list(100)
    return {"agents": agents, "selection_mode": deal.get("agent_selection_mode", "manual")}


# ============ LEAD-DEAL LINKAGES ============
@router.post("/lead-deal-linkages")
async def create_linkage(body: LinkageCreate, user: dict = Depends(require_roles("admin", "agent"))):
    deal = await db.deals.find_one({"id": body.deal_id}, {"_id": 0})
    if not deal:
        raise HTTPException(404, "Deal not found")
    existing = await db.lead_deal_linkages.find_one(
        {"lead_id": body.lead_id, "deal_id": body.deal_id}, {"_id": 0}
    )
    if existing:
        await db.lead_deal_linkages.update_one(
            {"id": existing["id"]},
            {"$set": {"pipeline_status": body.pipeline_status, "notes": body.notes,
                      "updated_at": _now()}},
        )
        return await db.lead_deal_linkages.find_one({"id": existing["id"]}, {"_id": 0})
    doc = {
        "id": _id(), "lead_id": body.lead_id, "deal_id": body.deal_id,
        "pipeline_status": body.pipeline_status, "notes": body.notes,
        "deal_title": deal.get("title"), "deal_value": deal.get("value", 0),
        "deal_expected_close_date": deal.get("expected_close_date"),
        "user_id": user["id"], "created_at": _now(), "updated_at": _now(),
    }
    await db.lead_deal_linkages.insert_one(doc)
    return _strip_id(doc)


@router.get("/lead-deal-linkages")
async def list_linkages(lead_id: Optional[str] = None, deal_id: Optional[str] = None,
                        _: dict = Depends(require_roles("admin", "agent"))):
    q = {}
    if lead_id:
        q["lead_id"] = lead_id
    if deal_id:
        q["deal_id"] = deal_id
    return await db.lead_deal_linkages.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


# ============ TASKS ============
@router.post("/tasks")
async def create_task(body: TaskCreate, user: dict = Depends(require_roles("admin", "agent"))):
    doc = body.model_dump()
    # enrich with denormalized labels
    if doc.get("lead_id"):
        ld = await db.leads.find_one({"id": doc["lead_id"]}, {"_id": 0})
        if ld:
            doc["company_name"] = ld.get("name") or ld.get("company")
            doc["lead_name"] = ld.get("name")
            doc["pic_name"] = doc.get("pic_name") or ld.get("pic_name") or ""
    if doc.get("deal_id"):
        dl = await db.deals.find_one({"id": doc["deal_id"]}, {"_id": 0})
        if dl:
            doc["deal_name"] = dl.get("title")
    if doc.get("assigned_to"):
        u = await db.users.find_one({"id": doc["assigned_to"]}, {"_id": 0})
        if u:
            doc["assigned_to_name"] = u.get("full_name") or u.get("email")
    doc.update({"id": _id(), "owner_id": user["id"], "created_at": _now(), "updated_at": _now()})
    await db.tasks.insert_one(doc)
    return _strip_id(doc)


@router.get("/tasks")
async def list_tasks(
    lead_id: Optional[str] = None,
    deal_id: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=200),
    _: dict = Depends(require_roles("admin", "agent")),
):
    q: Dict[str, Any] = {}
    if lead_id:
        q["lead_id"] = lead_id
    if deal_id:
        q["deal_id"] = deal_id
    if search:
        q["$or"] = [
            {"title": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}},
            {"company_name": {"$regex": search, "$options": "i"}},
            {"pic_name": {"$regex": search, "$options": "i"}},
        ]
    total = await db.tasks.count_documents(q)
    items = await (
        db.tasks.find(q, {"_id": 0})
        .sort("created_at", -1)
        .skip((page - 1) * limit)
        .limit(limit)
        .to_list(limit)
    )
    return {
        "items": items, "total": total, "page": page, "limit": limit,
        "total_pages": math.ceil(total / limit) if total else 0,
    }


@router.put("/tasks/{task_id}")
async def update_task(task_id: str, payload: dict, _: dict = Depends(require_roles("admin", "agent"))):
    allowed = ("title", "description", "lead_id", "deal_id", "status", "payment_status",
               "priority", "due_date", "assigned_to", "pic_name", "pipeline_status")
    updates = {k: v for k, v in payload.items() if k in allowed}
    if "lead_id" in updates and updates["lead_id"]:
        ld = await db.leads.find_one({"id": updates["lead_id"]}, {"_id": 0})
        if ld:
            updates["company_name"] = ld.get("name") or ld.get("company")
            updates["lead_name"] = ld.get("name")
    if "deal_id" in updates and updates["deal_id"]:
        dl = await db.deals.find_one({"id": updates["deal_id"]}, {"_id": 0})
        if dl:
            updates["deal_name"] = dl.get("title")
    if "assigned_to" in updates and updates["assigned_to"]:
        u = await db.users.find_one({"id": updates["assigned_to"]}, {"_id": 0})
        if u:
            updates["assigned_to_name"] = u.get("full_name") or u.get("email")
    updates["updated_at"] = _now()
    await db.tasks.update_one({"id": task_id}, {"$set": updates})
    return await db.tasks.find_one({"id": task_id}, {"_id": 0})


@router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, _: dict = Depends(require_roles("admin", "agent"))):
    await db.tasks.delete_one({"id": task_id})
    return {"deleted": True}


# ============ USERS (lookup for task assignment) ============
@router.get("/users")
async def list_users(
    page: int = Query(1, ge=1), limit: int = Query(50, ge=1, le=200),
    _: dict = Depends(require_roles("admin", "agent")),
):
    total = await db.users.count_documents({})
    users = await (
        db.users.find({}, {"_id": 0, "id": 1, "full_name": 1, "email": 1, "role": 1, "phone": 1})
        .sort("full_name", 1).skip((page - 1) * limit).limit(limit).to_list(limit)
    )
    return {"items": users, "total": total, "page": page, "limit": limit}


# ============ GOOGLE CALENDAR (stub — returns descriptive error if not configured) ============
@router.post("/google-calendar/sync-task/{task_id}")
async def sync_task_to_calendar(task_id: str, user: dict = Depends(require_roles("admin", "agent"))):
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(404, "Task not found")
    settings = await db.settings.find_one({"id": "app_settings"}, {"_id": 0}) or {}
    if not settings.get("google_oauth_client_id") or not settings.get("google_oauth_client_secret"):
        raise HTTPException(400, "Google Calendar not configured. Add Google OAuth keys in Settings.")
    # Placeholder — record the attempt
    await db.tasks.update_one({"id": task_id}, {"$set": {
        "calendar_synced": True, "calendar_synced_at": _now(),
        "updated_at": _now(),
    }})
    return {"synced": True, "message": "Calendar sync recorded (full OAuth flow coming soon)."}


# ============ AI AGENTS ============
@router.get("/ai-agents")
async def list_ai_agents(_: dict = Depends(require_roles("admin", "agent"))):
    agents = await db.ai_agents.find({}, {"_id": 0}).to_list(100)
    return {"agents": agents}


@router.post("/ai-agents")
async def create_ai_agent(body: AiAgentCreate, _: dict = Depends(require_roles("admin"))):
    doc = body.model_dump()
    doc.update({"id": _id(), "created_at": _now()})
    await db.ai_agents.insert_one(doc)
    return _strip_id(doc)


@router.delete("/ai-agents/{agent_id}")
async def delete_ai_agent(agent_id: str, _: dict = Depends(require_roles("admin"))):
    await db.ai_agents.delete_one({"id": agent_id})
    return {"deleted": True}


# ============ AI CALLS ============
@router.post("/ai-calls/initiate")
async def initiate_ai_call(body: AiCallInitiate, user: dict = Depends(require_roles("admin", "agent"))):
    lead = await db.leads.find_one({"id": body.lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    deal = await db.deals.find_one({"id": body.deal_id}, {"_id": 0})
    if not deal:
        raise HTTPException(404, "Deal not found")

    agent = None
    if body.agent_id:
        agent = await db.ai_agents.find_one({"id": body.agent_id}, {"_id": 0})
    elif body.agent_name:
        agent = await db.ai_agents.find_one({"name": body.agent_name}, {"_id": 0})
    if not agent:
        # Allow placeholder run when no agent — will fail at ElevenLabs but record the attempt
        agent = {"name": body.agent_name or "AI Agent", "agent_id": None}

    phone = body.phone or lead.get("phone")
    if not phone:
        raise HTTPException(400, "No phone number for this lead")

    result = await initiate_elevenlabs_call(
        phone=phone, lead=lead, deal=deal, agent=agent,
        knowledge=deal.get("knowledge_base_content"),
        purpose=body.call_purpose or "follow_up",
    )

    call = {
        "id": _id(),
        "lead_id": body.lead_id, "deal_id": body.deal_id,
        "agent_name": agent.get("name"), "agent_id": agent.get("id") or agent.get("agent_id"),
        "phone": phone, "lead_name": lead.get("name") or lead.get("pic_name"),
        "deal_title": deal.get("title"),
        "status": "initiated" if result.get("success") else "failed",
        "direction": "outbound", "source": "AI Call",
        "duration": None,
        "summary": f'AI call to discuss {deal.get("title")}. Agent: {agent.get("name")}',
        "recording_url": None, "transcript": None,
        "conversation_id": result.get("conversation_id"),
        "call_sid": result.get("call_sid"),
        "call_purpose": body.call_purpose,
        "error": result.get("error"),
        "initiated_by": user["id"], "created_at": _now(), "updated_at": _now(),
    }
    await db.ai_calls.insert_one(call)
    await db.activities.insert_one({
        "id": _id(), "entity_type": "lead", "entity_id": body.lead_id,
        "type": "ai_call", "action": "ai_call",
        "description": f"AI Call initiated to {phone} - {deal.get('title')}",
        "notes": result.get("error", "") or "Call started",
        "metadata": {"agent_name": agent.get("name"), "deal_title": deal.get("title"), "phone": phone},
        "call_id": call["id"],
        "user_id": user["id"], "user_name": user.get("full_name", ""),
        "created_at": _now(),
    })
    return {"success": result.get("success", False), "call": _strip_id(call),
            "message": result.get("error") or "AI call initiated"}


@router.get("/ai-calls/lead/{lead_id}")
async def list_lead_calls(lead_id: str, _: dict = Depends(require_roles("admin", "agent"))):
    calls = await db.ai_calls.find({"lead_id": lead_id}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return {"calls": calls}


@router.get("/ai-calls/{call_id}/details")
async def call_details(call_id: str, _: dict = Depends(require_roles("admin", "agent"))):
    call = await db.ai_calls.find_one({"id": call_id}, {"_id": 0})
    if not call:
        raise HTTPException(404, "Call not found")
    el = None
    if call.get("conversation_id"):
        el = await get_elevenlabs_conversation(call["conversation_id"])
        if el:
            await db.ai_calls.update_one({"id": call_id}, {"$set": {
                "transcript": el.get("transcript") or call.get("transcript"),
                "duration": el.get("duration_seconds") or call.get("duration"),
                "call_status": el.get("status"),
                "updated_at": _now(),
            }})
            call.update({"transcript": el.get("transcript") or call.get("transcript"),
                         "duration": el.get("duration_seconds")})
    return {
        "call": call,
        "elevenlabs_details": el,
        "has_audio": bool(el and el.get("has_audio")),
        "has_transcript": bool(call.get("transcript")),
        "audio_url": f"/api/ai-calls/{call_id}/audio" if (el and el.get("has_audio")) else None,
    }


@router.get("/ai-calls/{call_id}/audio")
async def call_audio(call_id: str, _: dict = Depends(require_roles("admin", "agent"))):
    call = await db.ai_calls.find_one({"id": call_id}, {"_id": 0})
    if not call or not call.get("conversation_id"):
        raise HTTPException(404, "No recording available")
    audio = await download_elevenlabs_audio(call["conversation_id"])
    if not audio:
        raise HTTPException(404, "Audio fetch failed")
    return StreamingResponse(io.BytesIO(audio), media_type="audio/mpeg")


@router.put("/ai-calls/{call_id}/interest")
async def set_interest(call_id: str, payload: dict, _: dict = Depends(require_roles("admin", "agent"))):
    await db.ai_calls.update_one(
        {"id": call_id},
        {"$set": {"interest_level": payload.get("interest_level"), "updated_at": _now()}},
    )
    return {"updated": True}


# ============ WHATSAPP ============
@router.get("/whatsapp/messages/{lead_id}")
async def list_whatsapp(lead_id: str, _: dict = Depends(require_roles("admin", "agent"))):
    items = await db.whatsapp_messages.find({"lead_id": lead_id}, {"_id": 0}).sort("created_at", 1).to_list(500)
    return {"messages": items}


@router.post("/whatsapp/send")
async def send_whatsapp_msg(body: WhatsAppSend, user: dict = Depends(require_roles("admin", "agent"))):
    if not body.phone or not body.message:
        raise HTTPException(400, "Phone and message required")
    result = await send_whatsapp(body.phone, body.message)
    msg = {
        "id": _id(),
        "lead_id": body.contact_id,
        "phone": body.phone,
        "message": body.message,
        "text": body.message,
        "sender": "me",
        "direction": "outbound",
        "status": "sent" if result.get("success") else "failed",
        "twilio_sid": result.get("sid"),
        "error": result.get("error"),
        "user_id": user["id"],
        "created_at": _now(),
        "timestamp": _now(),
    }
    await db.whatsapp_messages.insert_one(msg)
    if body.contact_id:
        await db.activities.insert_one({
            "id": _id(), "entity_type": "lead", "entity_id": body.contact_id,
            "type": "whatsapp", "action": "whatsapp",
            "description": "WhatsApp message sent",
            "notes": body.message[:200],
            "user_id": user["id"], "user_name": user.get("full_name", ""),
            "created_at": _now(),
        })
    if not result.get("success"):
        # Still return 200 — frontend renders failure state from db
        return _strip_id(msg)
    return _strip_id(msg)


# ============ MEETINGS ============
@router.get("/meetings/lead/{lead_id}")
async def list_meetings(lead_id: str, _: dict = Depends(require_roles("admin", "agent"))):
    items = await db.meetings.find({"lead_id": lead_id}, {"_id": 0}).sort("scheduled_at", -1).to_list(200)
    return {"meetings": items}


@router.post("/meetings/schedule")
async def schedule_meeting(body: MeetingSchedule, user: dict = Depends(require_roles("admin", "agent"))):
    lead = await db.leads.find_one({"id": body.lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    try:
        start = datetime.strptime(f"{body.date} {body.start_time}", "%Y-%m-%d %H:%M").replace(tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(400, "Invalid date/time")
    meeting = {
        "id": _id(),
        "lead_id": body.lead_id,
        "title": body.title,
        "description": body.description or "",
        "meeting_type": body.meeting_type,
        "scheduled_at": start.isoformat(),
        "duration_minutes": body.duration_minutes,
        "location": body.location or "",
        "send_invite": body.send_invite,
        "status": "scheduled",
        "user_id": user["id"],
        "created_at": _now(),
    }
    invite_sent = False
    invite_to = None
    if body.send_invite and lead.get("email"):
        ics = build_ics(meeting["id"], body.title, body.description or "",
                        start, body.duration_minutes, body.location or "")
        html = (
            f"<h2>{body.title}</h2>"
            f"<p>{body.description or ''}</p>"
            f"<p><b>When:</b> {start.strftime('%a %d %b %Y %H:%M UTC')}<br>"
            f"<b>Duration:</b> {body.duration_minutes} min<br>"
            f"<b>Type:</b> {body.meeting_type}"
            f"{f'<br><b>Location:</b> {body.location}' if body.location else ''}</p>"
            f"<p>— Insurance Tech CRM</p>"
        )
        r = await send_meeting_invite_email(lead["email"], body.title, html, ics)
        invite_sent = r.get("success", False)
        invite_to = lead["email"]
        meeting["invite_result"] = r
    await db.meetings.insert_one(meeting)
    await db.activities.insert_one({
        "id": _id(), "entity_type": "lead", "entity_id": body.lead_id,
        "type": "meeting", "action": "meeting",
        "description": f"Meeting scheduled: {body.title}",
        "notes": body.description or "",
        "scheduled_at": start.isoformat(),
        "user_id": user["id"], "user_name": user.get("full_name", ""),
        "created_at": _now(),
    })
    return {"meeting": _strip_id(meeting), "invite_sent": invite_sent, "invite_sent_to": invite_to}


# ============ LOOKUP ============
@router.get("/lookup/states")
async def list_states(_: dict = Depends(require_roles("admin", "agent"))):
    rows = await db.leads.distinct("state")
    states = sorted([s for s in rows if s])
    return {"states": states}


@router.get("/lookup/companies")
async def list_companies(_: dict = Depends(require_roles("admin", "agent"))):
    leads = await db.leads.find({}, {"_id": 0, "id": 1, "name": 1, "company": 1, "pic_name": 1, "state": 1}).to_list(2000)
    return {"companies": [{"id": ld["id"], "name": ld.get("company") or ld.get("name"),
                            "pic_name": ld.get("pic_name", ""), "state": ld.get("state", "")} for ld in leads]}


# ============ CSV EXPORT ============
@router.get("/leads/export/csv")
async def export_leads_csv(_: dict = Depends(require_roles("admin", "agent"))):
    import csv
    leads = await db.leads.find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    buf = io.StringIO()
    cols = ["name", "pic_name", "title", "email", "phone", "office_number", "ic_number",
            "passport_number", "country", "state", "city", "postcode", "address",
            "industry", "company_size", "website", "source", "pipeline_status", "status",
            "ai_score", "owner_name", "notes", "created_at"]
    writer = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
    writer.writeheader()
    for ld in leads:
        writer.writerow(ld)
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=leads_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"},
    )


# ============ LINKAGE PUT/DELETE (for Pipeline page) ============
@router.put("/lead-deal-linkages/{linkage_id}")
async def update_linkage(linkage_id: str, payload: dict, _: dict = Depends(require_roles("admin", "agent"))):
    updates = {k: v for k, v in payload.items() if k in ("pipeline_status", "notes")}
    updates["updated_at"] = _now()
    await db.lead_deal_linkages.update_one({"id": linkage_id}, {"$set": updates})
    return await db.lead_deal_linkages.find_one({"id": linkage_id}, {"_id": 0})


@router.delete("/lead-deal-linkages/{linkage_id}")
async def delete_linkage(linkage_id: str, _: dict = Depends(require_roles("admin", "agent"))):
    await db.lead_deal_linkages.delete_one({"id": linkage_id})
    return {"deleted": True}


# ============ DEAL KNOWLEDGE BASE ============
@router.post("/deals/{deal_id}/knowledge-base/upload")
async def upload_kb(deal_id: str, file: UploadFile = File(...), _: dict = Depends(require_roles("admin", "agent"))):
    content = await file.read()
    try:
        text = content.decode("utf-8", errors="ignore")
    except Exception:
        text = ""
    await db.deals.update_one({"id": deal_id}, {"$set": {
        "knowledge_base_filename": file.filename,
        "knowledge_base_content": text[:50000],
        "updated_at": _now(),
    }})
    return {"uploaded": True, "filename": file.filename, "size": len(content)}
